import sys, os, json, time, re, sqlite3
from datetime import datetime
from pathlib import Path
from flask import request, jsonify, g

# These are imported from sibling modules at registration time
# They will be set by server.py
db = None
auth = None
utils = None
obs_bucket_default = 'obs://openloong-zhengzhou-apps-private/data-collector-svc/align'

def register_routes(app, _db, _auth, _utils):
    """Register all API routes on the Flask app."""
    global db, auth, utils
    db = _db
    auth = _auth
    utils = _utils
    
    # Re-export for convenience
    get_db = db.get_db
    require_auth = auth.require_auth
    hash_password = auth.hash_password
    make_token = auth.make_token
    add_log = utils.add_log
    get_recent_logs = utils.get_recent_logs
    agent_req = utils.agent_req
    discover_task = utils.discover_task
    greedy_duration_select = utils.greedy_duration_select
    estimate_size = utils.estimate_size
    auto_assign_disks = utils.auto_assign_disks
    cache_upload = utils.cache_upload
    load_cached_file = utils.load_cached_file
    cleanup_cache = utils.cleanup_cache
    _parse_sheet_data = utils._parse_sheet_data
    sync_all_disks = utils.sync_all_disks

    # ═════ Auth ═════
    @app.route('/api/v1/auth/init', methods=['POST'])
    def auth_init():
        conn = get_db()
        count = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
        if count > 0:
            return jsonify({'ok': False, 'error': '已有用户，请登录'}), 400
        body = request.json or {}
        username = body.get('username', 'admin')
        pw = body.get('password', 'admin')
        conn.execute('INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)',
                   (username, hash_password(pw), 'admin'))
        conn.commit()
        add_log('info', f'管理员 {username} 登录', conn)
        return jsonify({'ok': True, 'message': f'管理员 {username} 创建成功'})

    @app.route('/api/v1/auth/login', methods=['POST'])
    def auth_login():
        body = request.json or {}
        conn = get_db()
        user = conn.execute('SELECT * FROM users WHERE username=?', (body.get('username', ''),)).fetchone()
        if not user or user['password_hash'] != hash_password(body.get('password', '')):
            return jsonify({'ok': False, 'error': '用户名或密码错误'}), 401
        token = make_token(user['id'], user['username'], user['role'])
        add_log('info', f'用户 {user["username"]} 登录', conn)
        return jsonify({'ok': True, 'token': token, 'username': user['username'], 'role': user['role']})

    @app.route('/api/v1/auth/me')
    @require_auth
    def auth_me():
        return jsonify({'ok': True, 'user': g.user})

    # ═════ Dashboard ═════
    @app.route('/api/v1/dashboard')
    @require_auth
    def dashboard():
        conn = get_db()
        agents = conn.execute('SELECT id, name, is_online FROM agents').fetchall()
        agent_count = len(agents)
        online_count = sum(1 for a in agents if a['is_online'])
        jobs = conn.execute('SELECT COUNT(*) as total, '
                          "SUM(CASE WHEN status='running' THEN 1 ELSE 0 END) as running, "
                          "SUM(CASE WHEN status='completed' THEN 1 ELSE 0 END) as done "
                          'FROM jobs').fetchone()
        total_bytes = conn.execute("SELECT COALESCE(SUM(total_bytes),0) FROM disks").fetchone()[0]
        disk_warnings = []
        for d in conn.execute('SELECT d.*, a.name as agent_name FROM disks d JOIN agents a ON a.id=d.agent_id ORDER BY d.free_bytes').fetchall():
            if d['total_bytes'] > 0:
                pct = (d['total_bytes'] - d['free_bytes']) / d['total_bytes'] * 100
                if pct > 90 or d['free_bytes'] < 50 * 1024**3:
                    disk_warnings.append({
                        'agent_name': d['agent_name'], 'label': d['label'],
                        'free_gb': round(d['free_bytes'] / (1024**3), 1), 'pct': round(pct, 1),
                    })
        return jsonify({
            'ok': True, 'agents': agent_count, 'online': online_count,
            'jobs_total': jobs['total'] or 0, 'jobs_running': jobs['running'] or 0,
            'jobs_completed': jobs['done'] or 0, 'total_disk_bytes': total_bytes,
            'disk_warnings': disk_warnings,
        })

    # ═════ Agents ═════
    @app.route('/api/v1/agents', methods=['GET'])
    @require_auth
    def list_agents():
        conn = get_db()
        agents = conn.execute('SELECT * FROM agents ORDER BY id').fetchall()
        result = []
        for a in agents:
            disks = conn.execute('SELECT * FROM disks WHERE agent_id=?', (a['id'],)).fetchall()
            result.append({
                'id': a['id'], 'name': a['name'], 'host': a['host'], 'port': a['port'],
                'is_online': bool(a['is_online']), 'last_seen': a['last_seen'],
                'disks': [dict(d) for d in disks],
            })
        return jsonify({'ok': True, 'agents': result})

    @app.route('/api/v1/agents', methods=['POST'])
    @require_auth
    def add_agent():
        body = request.json or {}
        name = body.get('name', '')
        host = body.get('host', '')
        port = body.get('port', 8081)
        api_key = body.get('api_key', '')
        if not name or not host:
            return jsonify({'ok': False, 'error': '缺少 name 或 host'}), 400
        conn = get_db()
        cur = conn.execute('INSERT INTO agents (name, host, port, api_key) VALUES (?,?,?,?)',
                         (name, host, port, api_key))
        conn.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid})

    @app.route('/api/v1/agents/<int:aid>', methods=['DELETE'])
    @require_auth
    def delete_agent(aid):
        conn = get_db()
        conn.execute('DELETE FROM disks WHERE agent_id=?', (aid,))
        conn.execute('DELETE FROM agents WHERE id=?', (aid,))
        conn.commit()
        return jsonify({'ok': True})

    @app.route('/api/v1/agents/<int:aid>/ping', methods=['POST'])
    @require_auth
    def ping_agent(aid):
        conn = get_db()
        agent = conn.execute('SELECT * FROM agents WHERE id=?', (aid,)).fetchone()
        if not agent:
            return jsonify({'ok': False, 'error': 'Agent not found'}), 404
        resp = agent_req(dict(agent), 'GET', '/api/v1/status')
        is_online = resp.get('ok', False)
        conn.execute('UPDATE agents SET is_online=?, last_seen=? WHERE id=?',
                   (int(is_online), datetime.now().isoformat(), aid))
        if is_online:
            disk_resp = agent_req(dict(agent), 'POST', '/api/v1/disks/scan', {'paths': []})
            if disk_resp.get('ok'):
                conn.execute('DELETE FROM disks WHERE agent_id=?', (aid,))
                for d in disk_resp.get('disks', []):
                    conn.execute('INSERT INTO disks (agent_id, path, label, total_bytes, free_bytes) VALUES (?,?,?,?,?)',
                               (aid, d['path'], d['label'], d['total_bytes'], d['free_bytes']))
        conn.commit()
        return jsonify({'ok': is_online, 'agent_status': resp})

    @app.route('/api/v1/agents/sync-disks', methods=['POST'])
    @require_auth
    def sync_all_disks_route():
        count = sync_all_disks()
        return jsonify({'ok': True, 'synced': count})

    # ═════ OBS Sources ═════
    @app.route('/api/v1/obs-sources', methods=['GET'])
    @require_auth
    def list_obs():
        conn = get_db()
        return jsonify({'ok': True, 'sources': [dict(r) for r in conn.execute('SELECT * FROM obs_sources').fetchall()]})

    @app.route('/api/v1/obs-sources', methods=['POST'])
    @require_auth
    def add_obs():
        body = request.json or {}
        conn = get_db()
        cur = conn.execute('INSERT INTO obs_sources (name, bucket, audit_bucket, binary_path) VALUES (?,?,?,?)',
                         (body['name'], body['bucket'], body.get('audit_bucket', ''), body.get('binary_path', 'obsutil')))
        conn.commit()
        return jsonify({'ok': True, 'id': cur.lastrowid})

    # ═════ Task Info ═════
    @app.route('/api/v1/tasks/info', methods=['POST'])
    @require_auth
    def task_info():
        body = request.json or {}
        task_ids = body.get('task_ids', [])
        durations = body.get('durations', {})
        obs_source_id = body.get('obs_source_id')
        conn = get_db()
        audit_bucket = ''
        if obs_source_id:
            src = conn.execute('SELECT * FROM obs_sources WHERE id=?', (obs_source_id,)).fetchone()
            if src:
                audit_bucket = src['audit_bucket'] or ''
        if not task_ids and 'task_id' in body:
            task_ids = [body['task_id']]
        results = []
        for tid in task_ids:
            eps = discover_task(tid, audit_bucket)
            selected = greedy_duration_select(eps, durations.get(tid, 0))
            size = estimate_size(selected)
            results.append({
                'task_id': tid, 'duration_h': durations.get(tid, 0),
                'episodes': [e['episode_id'] for e in selected],
                'episode_count': len(selected), 'total_bytes': size,
                'total_size_gb': round(size / (1024**3), 2),
            })
        return jsonify({'ok': True, 'tasks': results, 'episodes': results[0]['episodes'] if results else []})

    # ═════ Import ═════
    @app.route('/api/v1/import/parse', methods=['POST'])
    @require_auth
    def import_parse():
        if 'file' not in request.files:
            return jsonify({'ok': False, 'error': '请上传文件'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'ok': False, 'error': '空文件名'}), 400
        ids = set()
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        if ext in ('xlsx',):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(f, read_only=True)
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        for cell in row:
                            if cell and isinstance(cell, str):
                                for m in re.finditer(r'[0-9a-f]{32}', cell, re.I):
                                    ids.add(m.group().lower())
            except ImportError:
                return jsonify({'ok': False, 'error': '缺少 openpyxl'}), 500
        elif ext in ('xls',):
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=f.read())
                for sheet in wb.sheets():
                    for row in range(sheet.nrows):
                        for cell in sheet.row_values(row):
                            if cell and isinstance(cell, str):
                                for m in re.finditer(r'[0-9a-f]{32}', cell, re.I):
                                    ids.add(m.group().lower())
            except ImportError:
                return jsonify({'ok': False, 'error': '缺少 xlrd'}), 500
        else:
            content = f.read().decode('utf-8', errors='ignore')
            for m in re.finditer(r'[0-9a-f]{32}', content, re.I):
                ids.add(m.group().lower())
        ids = sorted(ids)
        if not ids:
            return jsonify({'ok': False, 'error': '未找到 32 位十六进制 task_id'}), 400
        return jsonify({'ok': True, 'ids': ids, 'count': len(ids)})

    @app.route('/api/v1/import/preview', methods=['POST'])
    @require_auth
    def import_preview():
        if 'file' not in request.files:
            return jsonify({'ok': False, 'error': '请上传文件'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'ok': False, 'error': '空文件名'}), 400
        ext = f.filename.rsplit('.', 1)[-1].lower() if '.' in f.filename else ''
        file_key = cache_upload(f)
        file_path = load_cached_file(file_key)
        sheets = []
        try:
            if ext in ('xlsx',):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                for ws in wb.worksheets:
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows: continue
                    si = _parse_sheet_data(rows, ws.max_row)
                    si['name'] = ws.title
                    sheets.append(si)
                wb.close()
            elif ext in ('xls',):
                import xlrd
                wb = xlrd.open_workbook(file_path)
                for sheet in wb.sheets():
                    rows = [sheet.row_values(i) for i in range(min(sheet.nrows, 500))]
                    if not rows: continue
                    si = _parse_sheet_data(rows, sheet.nrows)
                    si['name'] = sheet.name
                    sheets.append(si)
            else:
                content = open(file_path, encoding='utf-8', errors='ignore').read()
                lines = [l.split(',') if ',' in l else l.split('\t') if '\t' in l else [l]
                         for l in content.strip().split('\n') if l.strip()]
                if lines:
                    si = _parse_sheet_data(lines, len(lines))
                    si['name'] = 'Sheet1'
                    sheets.append(si)
        except Exception as e:
            return jsonify({'ok': False, 'error': f'解析失败: {str(e)}'}), 500
        if not sheets:
            return jsonify({'ok': False, 'error': '未找到有效数据'}), 400
        return jsonify({'ok': True, 'file_key': file_key, 'sheets': sheets})

    @app.route('/api/v1/import/extract', methods=['POST'])
    @require_auth
    def import_extract():
        body = request.json or {}
        file_key = body.get('file_key', '')
        sheet_mappings = body.get('sheets', [])
        if not file_key or not sheet_mappings:
            return jsonify({'ok': False, 'error': '缺少 file_key 或 sheets'}), 400
        file_path = load_cached_file(file_key)
        if not file_path:
            return jsonify({'ok': False, 'error': '文件已过期，请重新上传'}), 400
        ext = Path(file_path).suffix.lower().lstrip('.')
        tasks = []
        seen = set()
        try:
            if ext in ('xlsx',):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                for sm in sheet_mappings:
                    if sm.get('task_id_col') is None: continue
                    sheet_name = sm.get('name', '')
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        ws = wb.worksheets[int(sheet_name)] if sheet_name.isdigit() else wb.worksheets[0]
                    tid_col = sm.get('task_id_col')
                    dur_col = sm.get('duration_col')
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i == 0: continue
                        tid = str(row[tid_col]).strip() if tid_col is not None and tid_col < len(row) and row[tid_col] else ''
                        tid_lower = tid.lower()
                        if re.match(r'^[0-9a-f]{31,32}$', tid_lower) and tid_lower not in seen:
                            dur = 0
                            if dur_col is not None and dur_col < len(row):
                                try: dur = float(row[dur_col])
                                except: pass
                            tasks.append({'task_id': tid_lower, 'duration_h': dur})
                            seen.add(tid_lower)
                wb.close()
            elif ext in ('xls',):
                import xlrd
                wb = xlrd.open_workbook(file_path)
                for sm in sheet_mappings:
                    if sm.get('task_id_col') is None: continue
                    sheet_name = sm.get('name', '')
                    if sheet_name in wb.sheet_names():
                        sheet = wb.sheet_by_name(sheet_name)
                    else:
                        sheet = wb.sheet_by_index(int(sheet_name)) if sheet_name.isdigit() else wb.sheet_by_index(0)
                    tid_col = sm.get('task_id_col')
                    dur_col = sm.get('duration_col')
                    for i in range(1, sheet.nrows):
                        tid = str(sheet.cell_value(i, tid_col)).strip() if tid_col is not None and tid_col < sheet.ncols else ''
                        tid_lower = tid.lower()
                        if re.match(r'^[0-9a-f]{31,32}$', tid_lower) and tid_lower not in seen:
                            dur = 0
                            if dur_col is not None and dur_col < sheet.ncols:
                                try: dur = float(sheet.cell_value(i, dur_col))
                                except: pass
                            tasks.append({'task_id': tid_lower, 'duration_h': dur})
                            seen.add(tid_lower)
                wb.release_resources()
            else:
                content = open(file_path, encoding='utf-8', errors='ignore').read()
                for m in re.finditer(r'[0-9a-f]{31,32}', content, re.I):
                    tid_lower = m.group().lower()
                    if tid_lower not in seen:
                        tasks.append({'task_id': tid_lower, 'duration_h': 0})
                        seen.add(tid_lower)
        except Exception as e:
            import traceback; traceback.print_exc()
            return jsonify({'ok': False, 'error': f'提取失败: {str(e)}'}), 500
        if not tasks:
            skipped = sum(1 for sm in sheet_mappings if sm.get('task_id_col') is None)
            return jsonify({'ok': False, 'error': f'未找到有效的 task_id ({len(sheet_mappings)} sheet, {skipped} 个未选列)'}), 400
        return jsonify({'ok': True, 'tasks': tasks, 'total': len(tasks)})

    # ═════ Tasks (Assign, Prepare, Pending, Allocate) ═════
    @app.route('/api/v1/tasks/assign', methods=['POST'])
    @require_auth
    def task_assign():
        body = request.json or {}
        task_ids = body.get('task_ids', [])
        agent_id = body.get('agent_id')
        durations = body.get('durations', {})
        obs_source_id = body.get('obs_source_id')
        if not task_ids or not agent_id:
            return jsonify({'ok': False, 'error': '缺少 task_ids 或 agent_id'}), 400
        conn = get_db()
        audit_bucket = ''
        if obs_source_id:
            src = conn.execute('SELECT * FROM obs_sources WHERE id=?', (obs_source_id,)).fetchone()
            if src: audit_bucket = src['audit_bucket'] or ''
        task_data = []
        for tid in task_ids:
            eps = discover_task(tid, audit_bucket)
            selected = greedy_duration_select(eps, durations.get(tid, 0))
            task_data.append({
                'task_id': tid, 'duration_h': durations.get(tid, 0),
                'episodes': [e['episode_id'] for e in selected],
                'total_bytes': estimate_size(selected),
            })
        task_data.sort(key=lambda t: t['duration_h'], reverse=True)
        disks = conn.execute('SELECT * FROM disks WHERE agent_id=?', (agent_id,)).fetchall()
        if not disks:
            return jsonify({'ok': False, 'error': '该机器无磁盘信息，请先检测'}), 400
        disk_infos = [{'path': d['path'], 'label': d['label'], 'total': d['total_bytes'], 'free': d['free_bytes']} for d in disks]
        disk_infos.sort(key=lambda d: d['free'])
        assignments = []
        disk_free = {d['path']: d['free'] for d in disk_infos}
        for t in task_data:
            size = t['total_bytes']
            best = None
            for d in sorted(disk_infos, key=lambda x: disk_free[x['path']], reverse=True):
                if size <= disk_free[d['path']]:
                    best = d['path']; break
            if best is None:
                best = max(disk_infos, key=lambda x: disk_free[x['path']])['path']
            disk_free[best] -= size
            assignments.append({
                'task_id': t['task_id'], 'duration_h': t['duration_h'],
                'disk_path': best, 'episodes': t['episodes'], 'estimated_bytes': size,
            })
        return jsonify({'ok': True, 'assignments': assignments})

    @app.route('/api/v1/tasks/prepare', methods=['POST'])
    @require_auth
    def tasks_prepare():
        body = request.json or {}
        task_ids = body.get('task_ids', [])
        agent_id = body.get('agent_id')
        durations = body.get('durations', {})
        obs_source_id = body.get('obs_source_id')
        if not task_ids or not agent_id:
            return jsonify({'ok': False, 'error': '缺少 task_ids 或 agent_id'}), 400
        conn = get_db()
        audit_bucket = ''
        if obs_source_id:
            src = conn.execute('SELECT * FROM obs_sources WHERE id=?', (obs_source_id,)).fetchone()
            if src: audit_bucket = src['audit_bucket'] or ''
        task_data = []
        for tid in task_ids:
            eps = discover_task(tid, audit_bucket)
            selected = greedy_duration_select(eps, durations.get(tid, 0))
            task_data.append({
                'task_id': tid, 'duration_h': durations.get(tid, 0),
                'episodes': [e['episode_id'] for e in selected],
                'episode_count': len(selected), 'total_bytes': estimate_size(selected),
            })
        task_data.sort(key=lambda t: t['duration_h'], reverse=True)
        assigned = auto_assign_disks(conn, agent_id, task_data)
        if not assigned:
            return jsonify({'ok': False, 'error': '该机器无磁盘信息，请先检测'}), 400
        for t in assigned:
            conn.execute(
                'INSERT OR REPLACE INTO pending_tasks (task_id, agent_id, disk_path, episode_count, estimated_bytes, episodes, status) VALUES (?,?,?,?,?,?,?)',
                (t['task_id'], agent_id, t['disk_path'], t['episode_count'], t['total_bytes'],
                 ','.join(t['episodes'][:500]), 'pending')
            )
        conn.commit()
        return jsonify({'ok': True, 'tasks': assigned, 'count': len(assigned)})

    @app.route('/api/v1/tasks/pending', methods=['GET'])
    @require_auth
    def tasks_pending():
        conn = get_db()
        agent_id = request.args.get('agent_id', '')
        status = request.args.get('status', '')
        query = 'SELECT * FROM pending_tasks'
        params = []
        conditions = []
        if agent_id:
            conditions.append('agent_id=?'); params.append(agent_id)
        if status:
            conditions.append('status=?'); params.append(status)
        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)
        query += ' ORDER BY created_at DESC'
        rows = conn.execute(query, params).fetchall()
        return jsonify({'ok': True, 'tasks': [dict(r) for r in rows]})

    @app.route('/api/v1/tasks/pending/retry', methods=['POST'])
    @require_auth
    def tasks_pending_retry():
        body = request.json or {}
        ids = body.get('ids', [])
        if not ids:
            return jsonify({'ok': False, 'error': '缺少 ids'}), 400
        conn = get_db()
        for pid in ids:
            conn.execute("UPDATE pending_tasks SET status='pending' WHERE id=?", (pid,))
        conn.commit()
        return jsonify({'ok': True, 'count': len(ids)})

    @app.route('/api/v1/tasks/pending/<int:pid>', methods=['DELETE'])
    @require_auth
    def tasks_pending_delete(pid):
        conn = get_db()
        conn.execute('DELETE FROM pending_tasks WHERE id=?', (pid,))
        conn.commit()
        return jsonify({'ok': True})

    @app.route('/api/v1/tasks/allocate-preview', methods=['POST'])
    @require_auth
    def allocate_preview():
        body = request.json or {}
        agent_id = body.get('agent_id')
        if not agent_id:
            return jsonify({'ok': False, 'error': '缺少 agent_id'}), 400
        conn = get_db()
        pending = conn.execute("SELECT * FROM pending_tasks WHERE agent_id=? AND status='pending' ORDER BY estimated_bytes DESC",
                             (agent_id,)).fetchall()
        if not pending:
            return jsonify({'ok': False, 'error': '无待处理任务，请先导入'}), 400
        disks = conn.execute('SELECT * FROM disks WHERE agent_id=? ORDER BY free_bytes', (agent_id,)).fetchall()
        if not disks:
            return jsonify({'ok': False, 'error': '该机器无磁盘信息，请先检测'}), 400
        disk_free = {d['path']: d['free_bytes'] for d in disks}
        disk_label = {d['path']: d['label'] or os.path.basename(d['path']) for d in disks}
        disk_tasks = {d['path']: [] for d in disks}
        for t in pending:
            size = t['estimated_bytes'] or 0
            for d in disks:
                if size <= disk_free[d['path']]:
                    disk_tasks[d['path']].append({
                        'id': t['id'], 'task_id': t['task_id'],
                        'episode_count': t['episode_count'],
                        'size_gb': round(size / (1024**3), 2),
                    })
                    disk_free[d['path']] -= size
                    break
        plan = []
        for d in disks:
            assigned = disk_tasks[d['path']]
            if not assigned: continue
            plan.append({
                'disk_path': d['path'], 'label': disk_label[d['path']],
                'free_gb': round(d['free_bytes'] / (1024**3), 2),
                'task_count': len(assigned),
                'total_size_gb': round(sum(t['size_gb'] for t in assigned), 2),
                'remaining_gb': round(disk_free[d['path']] / (1024**3), 2),
                'tasks': assigned,
            })
        return jsonify({
            'ok': True, 'plan': plan,
            'unassigned': len(pending) - sum(p['task_count'] for p in plan),
            'total_tasks': len(pending),
        })

    # ═════ Jobs ═════
    @app.route('/api/v1/jobs', methods=['GET'])
    @require_auth
    def list_jobs():
        conn = get_db()
        jobs = conn.execute('SELECT * FROM jobs ORDER BY id DESC LIMIT 50').fetchall()
        return jsonify({'ok': True, 'jobs': [dict(j) for j in jobs]})

    @app.route('/api/v1/jobs/<int:jid>')
    @require_auth
    def get_job(jid):
        conn = get_db()
        job = conn.execute('SELECT * FROM jobs WHERE id=?', (jid,)).fetchone()
        if not job:
            return jsonify({'ok': False, 'error': 'Not found'}), 404
        tasks = conn.execute('SELECT * FROM job_tasks WHERE job_id=?', (jid,)).fetchall()
        return jsonify({'ok': True, 'job': dict(job), 'tasks': [dict(t) for t in tasks]})

    @app.route('/api/v1/jobs/<int:jid>/progress')
    @require_auth
    def job_progress(jid):
        conn = get_db()
        job = conn.execute('SELECT * FROM jobs WHERE id=?', (jid,)).fetchone()
        if not job:
            return jsonify({'ok': False, 'error': 'Not found'}), 404
        if not job['agent_job_id']:
            return jsonify({'ok': True, 'job_id': jid, 'completed': job['completed'],
                          'failed': job['failed'], 'total': job['total_episodes'], 'status': job['status']})
        if job['agent_id']:
            agent = conn.execute('SELECT * FROM agents WHERE id=?', (job['agent_id'],)).fetchone()
            if agent:
                resp = agent_req(dict(agent), 'GET', f'/api/v1/progress/{job["agent_job_id"]}')
                if resp.get('ok'):
                    completed = resp.get('completed', 0)
                    failed = resp.get('failed', 0)
                    total = resp.get('total', 0)
                    if total > 0 and (completed + failed) >= total:
                        new_status = 'completed' if failed == 0 else 'partial'
                        conn.execute('UPDATE jobs SET status=?, completed=?, failed=?, completed_at=? WHERE id=?',
                                   (new_status, completed, failed, datetime.now().isoformat(), jid))
                        conn.execute("UPDATE pending_tasks SET status='completed' WHERE agent_id=? AND status='running'",
                                   (job['agent_id'],))
                        conn.commit()
                    return jsonify(resp)
        return jsonify({'ok': True, 'completed': 0, 'failed': 0, 'total': 0})

    @app.route('/api/v1/jobs', methods=['POST'])
    @require_auth
    def create_job():
        body = request.json or {}
        name = body.get('name', f'job_{int(time.time())}')
        agent_id = body.get('agent_id')
        tasks_config = body.get('tasks', [])
        obs_source_id = body.get('obs_source_id')
        use_pending = body.get('use_pending', False)
        if not agent_id:
            return jsonify({'ok': False, 'error': '需要 agent_id'}), 400
        conn = get_db()
        agent = conn.execute('SELECT * FROM agents WHERE id=?', (agent_id,)).fetchone()
        if not agent:
            return jsonify({'ok': False, 'error': 'Agent not found'}), 404
        obs = conn.execute('SELECT * FROM obs_sources WHERE id=?', (obs_source_id,)).fetchone()
        bucket = obs['bucket'] if obs else obs_bucket_default
        audit_bucket = obs['audit_bucket'] if obs and obs['audit_bucket'] else ''
        if use_pending and not tasks_config:
            pending = conn.execute("SELECT * FROM pending_tasks WHERE agent_id=? AND status='pending'", (agent_id,)).fetchall()
            if not pending:
                return jsonify({'ok': False, 'error': '无待处理任务'}), 400
            tasks_config = [{'task_id': p['task_id'], 'disk_path': p['disk_path'],
                           'episodes': p['episodes'].split(',') if p['episodes'] else [],} for p in pending]
        if not tasks_config:
            return jsonify({'ok': False, 'error': '需要 tasks 或 use_pending=true'}), 400
        need_assign = any(not t.get('disk_path') for t in tasks_config)
        need_eps = any(not t.get('episodes') or len(t['episodes']) == 0 for t in tasks_config)
        if need_eps or need_assign:
            tmp_data = []
            for t in tasks_config:
                eps = t.get('episodes', [])
                if not eps or len(eps) == 0:
                    eps = discover_task(t['task_id'], audit_bucket)
                    selected = greedy_duration_select(eps, 0)
                    eps = [e['episode_id'] for e in selected]
                tmp_data.append({'task_id': t['task_id'], 'total_bytes': estimate_size([{'file_size': 0} for _ in eps]), 'disk_path': t.get('disk_path', '')})
                t['episodes'] = eps
            if need_assign:
                assigned = auto_assign_disks(conn, agent_id, tmp_data)
                if assigned:
                    for i, t in enumerate(tasks_config):
                        t['disk_path'] = assigned[i]['disk_path']
        total_eps = sum(len(t.get('episodes', [])) for t in tasks_config)
        cur = conn.execute('INSERT INTO jobs (name, status, total_episodes, obs_source_id, agent_id, created_by) VALUES (?,?,?,?,?,?)',
                         (name, 'running', total_eps, obs_source_id, agent_id, g.user['id']))
        job_id = cur.lastrowid
        for t in tasks_config:
            conn.execute('INSERT INTO job_tasks (job_id, task_id, disk_path, selected_episodes) VALUES (?,?,?,?)',
                       (job_id, t['task_id'], t.get('disk_path', ''), len(t.get('episodes', []))))
        conn.commit()
        agent_job_id = f'srv_job_{job_id}'
        payload = {'job_id': agent_job_id, 'tasks': [{'task_id': t['task_id'], 'episodes': t['episodes'], 'disk': t['disk_path']} for t in tasks_config], 'obs_source': bucket}
        resp = agent_req(dict(agent), 'POST', '/api/v1/start', payload)
        if resp.get('ok'):
            conn.execute('UPDATE jobs SET agent_job_id=? WHERE id=?', (agent_job_id, job_id))
            conn.execute("UPDATE pending_tasks SET status='running', disk_path='' WHERE agent_id=? AND status='pending'", (agent_id,))
            conn.commit()
        else:
            conn.execute('UPDATE jobs SET status=? WHERE id=?', ('failed', job_id))
            conn.commit()
            return jsonify({'ok': False, 'error': f'Agent 启动失败: {resp.get("error", "unknown")}'}), 500
        add_log('info', f'作业 #{job_id} 已创建 ({len(tasks_config)} task)', conn)
        return jsonify({'ok': True, 'job_id': job_id, 'agent_job_id': agent_job_id})

    @app.route('/api/v1/jobs/<int:jid>/stop', methods=['POST'])
    @require_auth
    def stop_job(jid):
        conn = get_db()
        job = conn.execute('SELECT * FROM jobs WHERE id=?', (jid,)).fetchone()
        if not job or not job['agent_job_id']:
            return jsonify({'ok': False, 'error': 'Not found'}), 404
        agent = conn.execute('SELECT * FROM agents WHERE id=?', (job['agent_id'],)).fetchone()
        if agent:
            agent_req(dict(agent), 'POST', f'/api/v1/stop/{job["agent_job_id"]}')
        conn.execute("UPDATE jobs SET status='stopped' WHERE id=?", (jid,))
        conn.commit()
        add_log('info', f'作业 #{jid} 已停止', conn)
        return jsonify({'ok': True})

    # ═════ Monitor ═════
    @app.route('/api/v1/logs/file')
    @require_auth
    def logs_file():
        from pathlib import Path
        LOG_FILE = Path(__file__).parent.parent / 'dwheel-server.log'
        lines = request.args.get('lines', 200, type=int)
        if not LOG_FILE.exists():
            return jsonify({'ok': True, 'lines': []})
        with open(str(LOG_FILE), encoding='utf-8') as f:
            content = f.readlines()
        return jsonify({'ok': True, 'lines': content[-lines:], 'total': len(content)})

    @app.route('/api/v1/monitor')
    @require_auth
    def monitor():
        conn = get_db()
        logs = get_recent_logs(conn, 30)
        jobs = conn.execute('SELECT * FROM jobs ORDER BY id DESC LIMIT 20').fetchall()
        job_list = []
        for j in jobs:
            jd = dict(j)
            if jd['agent_job_id'] and jd['agent_id']:
                agent = conn.execute('SELECT * FROM agents WHERE id=?', (jd['agent_id'],)).fetchone()
                if agent:
                    resp = agent_req(dict(agent), 'GET', f'/api/v1/progress/{jd["agent_job_id"]}')
                    if resp.get('ok'):
                        jd['agent_completed'] = resp.get('completed', 0)
                        jd['agent_failed'] = resp.get('failed', 0)
                        jd['agent_total'] = resp.get('total', 0)
                        jd['agent_elapsed'] = resp.get('elapsed', 0)
            job_list.append(jd)
        disks = conn.execute('SELECT d.*, a.name as agent_name FROM disks d JOIN agents a ON a.id=d.agent_id ORDER BY d.agent_id').fetchall()
        return jsonify({'ok': True, 'logs': logs, 'jobs': [dict(j) for j in job_list[:10]], 'disks': [dict(d) for d in disks]})
